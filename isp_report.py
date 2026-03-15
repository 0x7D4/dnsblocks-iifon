#!/usr/bin/env python3
"""
isp_report.py — ISP DNS Censorship Comparison Report Generator  v2
════════════════════════════════════════════════════════════════════
Interactively select ISPs, pull ALL data from PostgreSQL, and produce
a professional multi-sheet Excel report with:
  • Full summary statistics
  • Per-ISP detail: metadata, category breakdown, timing
  • ALL blocked domains with complete DNS query details
    (TTL · all IPs · CNAME chain · rcode · response time)
  • Cross-ISP overlap with embedded Venn diagram
  • Unique-blocks analysis

Usage:
    python isp_report.py
    python isp_report.py --db postgresql://user:pass@host/db
    python isp_report.py --output report.xlsx

Requirements:
    pip install psycopg2-binary openpyxl matplotlib python-dotenv
"""

import argparse
import io
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path


# ── Auto-install ───────────────────────────────────────────────────────────────
def _ensure(pkg, mod=None):
    m = mod or pkg
    try:
        return __import__(m)
    except ImportError:
        print(f"[INFO] Installing {pkg}...")
        subprocess.run([sys.executable, "-m", "pip", "install", pkg, "-q"],
                       capture_output=True)
        return __import__(m)

_ensure("psycopg2-binary", "psycopg2")
_ensure("openpyxl")
_ensure("matplotlib")

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")
    load_dotenv()
except ImportError:
    for _p in [Path(__file__).resolve().parent / ".env", Path(".env")]:
        if _p.exists():
            with open(_p) as _f:
                for _l in _f:
                    _l = _l.strip()
                    if _l and not _l.startswith("#") and "=" in _l:
                        k, v = _l.split("=", 1)
                        os.environ.setdefault(k.strip(), v.strip())
            break

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import Circle


# ══════════════════════════════════════════════════════════════════════════════
#  Styling helpers — clean, light, high-contrast
# ══════════════════════════════════════════════════════════════════════════════

# Palette — light backgrounds, dark crisp text
BG_HEADER_DARK   = "1E3A5F"   # deep navy — title rows only
BG_HEADER_MID    = "2563EB"   # blue — column headers
BG_HEADER_LIGHT  = "DBEAFE"   # pale blue — sub-headers
BG_ROW_ALT       = "F0F4FA"   # faint blue-grey — alternating rows
BG_WHITE         = "FFFFFF"
BG_RED_LIGHT     = "FEE2E2"   # blocked rows
BG_GREEN_LIGHT   = "DCFCE7"   # accessible rows
BG_AMBER_LIGHT   = "FEF9C3"   # changed rows
BG_GREY_LIGHT    = "F3F4F6"   # timeout rows
BG_SECTION       = "EFF6FF"   # section dividers

FG_WHITE         = "FFFFFF"
FG_DARK          = "111827"   # almost-black body text
FG_BLUE_DARK     = "1E3A5F"
FG_MUTED         = "6B7280"
FG_RED           = "991B1B"
FG_GREEN         = "166534"

STATUS_BG = {
    "blocked":          BG_RED_LIGHT,
    "blocked_nxdomain": BG_RED_LIGHT,
    "blocked_servfail": BG_RED_LIGHT,
    "accessible":       BG_GREEN_LIGHT,
    "changed":          BG_AMBER_LIGHT,
    "timeout":          BG_GREY_LIGHT,
}
STATUS_LABEL = {
    "blocked":          "Blocked — IP Redirect",
    "blocked_nxdomain": "Blocked — NXDOMAIN",
    "blocked_servfail": "Blocked — SERVFAIL",
    "accessible":       "Accessible",
    "changed":          "Changed IP",
    "timeout":          "Timeout / Error",
}
CATEGORY_NAMES = {
    "MOV":   "Movies / Streaming",
    "PORN":  "Adult Content",
    "FILE":  "File Sharing",
    "GMB":   "Gambling",
    "LIVE":  "Live TV / IPTV",
    "MISC":  "Miscellaneous",
    "IPTM":  "IPTV / Media",
    "MILX":  "Military / Extremism",
    "MAL":   "Malware / Phishing",
    "MUS":   "Music Piracy",
    "VISA":  "Visa / Immigration",
    "NEWS":  "News / Media",
    "COMM":  "Communications",
    "BIZ":   "Business",
    "ESC":   "Escort / Adult Services",
    "GRP":   "Groups / Forums",
    "POLR":  "Political",
    "HOST":  "Hosting / VPN",
    "GOVT":  "Government",
    "COMT":  "Community",
    "CULTR": "Culture / Arts",
    "COIN":  "Cryptocurrency",
    "REL":   "Religion",
    "HACK":  "Hacking / Security",
    "ANON":  "Anonymizers",
    "UNCAT": "Uncategorised",
}

def _side(style="thin"):
    return Side(border_style=style, color="CBD5E1")

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom_only():
    return Border(bottom=_side("medium"))

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def _font(bold=False, size=11, color="111827", italic=False):
    return Font(bold=bold, size=size, color=color.lstrip("#"), italic=italic, name="Calibri")

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _right():
    return Alignment(horizontal="right", vertical="center")

def title_cell(ws, row, col, value, end_col=None, size=14):
    if end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, size=size, color=FG_WHITE, name="Calibri")
    c.fill = _fill(BG_HEADER_DARK)
    c.alignment = _center()
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 26)
    return c

def header_cell(ws, row, col, value, end_col=None, size=11, bg=BG_HEADER_MID, fg=FG_WHITE, wrap=True):
    if end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, size=size, color=fg, name="Calibri")
    c.fill = _fill(bg)
    c.alignment = _center(wrap=wrap)
    c.border = _border()
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 30)
    return c

def section_cell(ws, row, col, value, end_col=None):
    if end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, size=11, color=BG_HEADER_DARK, name="Calibri")
    c.fill = _fill(BG_HEADER_LIGHT)
    c.alignment = _left()
    c.border = Border(bottom=_side("medium"), top=_side("medium"))
    ws.row_dimensions[row].height = 22
    return c

def data_cell(ws, row, col, value, bg=BG_WHITE, fg=FG_DARK,
              bold=False, fmt=None, align="center", wrap=False, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=fg, name="Calibri")
    c.fill = _fill(bg)
    if align == "left":
        c.alignment = _left(wrap=wrap)
    elif align == "right":
        c.alignment = _right()
    else:
        c.alignment = _center(wrap=wrap)
    c.border = _border()
    if fmt:
        c.number_format = fmt
    return c

def set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ══════════════════════════════════════════════════════════════════════════════
#  Venn diagram (matplotlib → PNG → embedded in sheet)
# ══════════════════════════════════════════════════════════════════════════════

def make_venn_png(isp_data, blocked_sets):
    """
    Draw a proportionally-scaled Venn diagram for 2 ISPs.
    Returns PNG bytes.
    """
    if len(isp_data) != 2:
        return None

    run_ids = [d["run_id"] for d in isp_data]
    set_a = blocked_sets.get(run_ids[0], set())
    set_b = blocked_sets.get(run_ids[1], set())
    a_total = len(set_a)
    b_total = len(set_b)
    both    = len(set_a & set_b)
    a_only  = a_total - both
    b_only  = b_total - both
    label_a = isp_data[0]["label"]
    label_b = isp_data[1]["label"]

    fig, ax = plt.subplots(figsize=(11, 6.5))
    fig.patch.set_facecolor("#F8FAFF")
    ax.set_facecolor("#F8FAFF")
    ax.set_xlim(0, 11)
    ax.set_ylim(0, 6.5)
    ax.axis("off")

    max_t = max(a_total, b_total, 1)
    r_a   = 2.4 * (a_total / max_t) ** 0.5
    r_b   = 2.4 * (b_total / max_t) ** 0.5
    r_b   = max(r_b, 0.65)

    # How much circles overlap: scale by overlap fraction of smaller set
    smaller = min(a_total, b_total) or 1
    overlap_frac = both / smaller
    # Distance between centres: r_a+r_b means no overlap; |r_a-r_b| means fully inside
    # Lerp between the two based on overlap
    d_max = r_a + r_b
    d_min = abs(r_a - r_b) + 0.05
    dist  = d_max - (d_max - d_min) * min(overlap_frac * 1.5, 0.92)

    cx_a, cx_b, cy = 5.5 - dist / 2, 5.5 + dist / 2, 3.2

    # Draw circles
    for cx, r, fc, ec in [(cx_a, r_a, "#3B82F6", "#1D4ED8"),
                           (cx_b, r_b, "#EF4444", "#B91C1C")]:
        ax.add_patch(Circle((cx, cy), r, color=fc, alpha=0.28, zorder=2))
        ax.add_patch(Circle((cx, cy), r, fill=False, edgecolor=ec, linewidth=2.5, zorder=3))

    # ISP name labels (above circles)
    ax.text(cx_a - r_a * 0.3, cy + r_a + 0.22, label_a,
            ha="center", va="bottom", fontsize=15, fontweight="bold",
            color="#1D4ED8", zorder=5)
    ax.text(cx_b + r_b * 0.3, cy + r_b + 0.22, label_b,
            ha="center", va="bottom", fontsize=15, fontweight="bold",
            color="#B91C1C", zorder=5)

    # Count labels inside regions
    left_x  = cx_a - dist * 0.25
    right_x = cx_b + dist * 0.25
    mid_x   = (cx_a + cx_b) / 2

    def count_label(x, y, n, label, color):
        ax.text(x, y + 0.22, f"{n:,}", ha="center", va="center",
                fontsize=13, fontweight="bold", color=color, zorder=6)
        ax.text(x, y - 0.28, label, ha="center", va="center",
                fontsize=9, color=color, zorder=6, style="italic")

    count_label(left_x,  cy, a_only, f"{label_a} only", "#1E40AF")
    count_label(mid_x,   cy, both,   "both ISPs",        "#6B21A8")
    count_label(right_x, cy, b_only, f"{label_b} only", "#991B1B")

    # Total badges below each circle
    for cx, r, total, ec in [(cx_a, r_a, a_total, "#1D4ED8"),
                               (cx_b, r_b, b_total, "#B91C1C")]:
        ax.text(cx, cy - r - 0.42, f"Total blocked: {total:,}",
                ha="center", va="top", fontsize=10, color="#374151",
                bbox=dict(boxstyle="round,pad=0.35", facecolor="white",
                          edgecolor=ec, linewidth=1.8, zorder=5))

    # Block rate annotation
    pct_a = a_total / (a_total + b_total) * 100 if (a_total + b_total) else 0
    ax.text(5.5, 0.25,
            f"Shared (both blocked): {both:,}  |  "
            f"Union (ever blocked): {a_total + b_total - both:,}  |  "
            f"Jaccard similarity: {both / (a_total + b_total - both) * 100:.1f}%"
            if (a_total + b_total - both) else "",
            ha="center", va="bottom", fontsize=9, color="#6B7280")

    ax.set_title("Blocked Domain Overlap — Venn Diagram",
                 fontsize=17, fontweight="bold", color="#1E3A5F", pad=12)

    buf = io.BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="#F8FAFF")
    plt.close(fig)
    buf.seek(0)
    return buf


def make_category_chart_png(categories, isp_label):
    """Horizontal bar chart of blocked categories. Returns PNG bytes."""
    if not categories:
        return None
    cats = [(CATEGORY_NAMES.get(c, c), n) for c, n in categories[:15]]
    cats.reverse()
    names = [c[0] for c in cats]
    counts = [c[1] for c in cats]
    total = sum(counts)

    fig, ax = plt.subplots(figsize=(9, max(4, len(cats) * 0.5 + 1.5)))
    fig.patch.set_facecolor("#F8FAFF")
    ax.set_facecolor("#F8FAFF")

    cmap = plt.cm.Blues
    colors = [cmap(0.45 + 0.5 * (i / max(len(cats)-1, 1))) for i in range(len(cats))]
    bars = ax.barh(names, counts, color=colors, edgecolor="white", linewidth=0.8)

    for bar, cnt in zip(bars, counts):
        pct = cnt / total * 100 if total else 0
        ax.text(bar.get_width() + max(counts) * 0.01, bar.get_y() + bar.get_height() / 2,
                f"{cnt:,}  ({pct:.1f}%)",
                va="center", ha="left", fontsize=9, color="#111827")

    ax.set_xlabel("Blocked Domain Count", fontsize=10)
    ax.set_title(f"Blocked Domains by Category — {isp_label}",
                 fontsize=12, fontweight="bold", color="#1E3A5F", pad=10)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(labelsize=9)
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="#F8FAFF")
    plt.close(fig)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
#  Database functions
# ══════════════════════════════════════════════════════════════════════════════

def db_connect(connstr):
    conn = psycopg2.connect(connstr)
    conn.autocommit = True
    return conn


def fetch_available_isps(conn):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT id, label, resolver_ip, started_at, total_domains, asn_org
            FROM measurement_runs
            WHERE completed_at IS NOT NULL
            ORDER BY label, started_at DESC
        """)
        return cur.fetchall()


def fetch_run_meta(conn, run_id):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT id, label, resolver_ip, started_at, completed_at,
                   total_domains, block_ip, asn, asn_org, country, public_ip
            FROM measurement_runs WHERE id = %s
        """, (run_id,))
        cols = [d[0] for d in cur.description]
        row  = cur.fetchone()
        return dict(zip(cols, row)) if row else {}


def fetch_run_summary(conn, run_id):
    with conn.cursor() as cur:
        cur.execute("SELECT status, count FROM run_summary WHERE run_id = %s", (run_id,))
        return {r[0]: r[1] for r in cur.fetchall()}


def fetch_category_breakdown(conn, run_id):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT category, blocked_count
            FROM category_breakdowns WHERE run_id = %s
            ORDER BY blocked_count DESC
        """, (run_id,))
        return cur.fetchall()


def fetch_avg_response_times(conn, run_id):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT query_type,
                   ROUND(AVG(response_time_ms)::numeric, 2),
                   ROUND(PERCENTILE_CONT(0.5)
                         WITHIN GROUP(ORDER BY response_time_ms)::numeric, 2)
            FROM dns_queries
            WHERE run_id = %s AND response_time_ms IS NOT NULL
            GROUP BY query_type
        """, (run_id,))
        return {r[0]: {"avg": float(r[1]), "median": float(r[2])} for r in cur.fetchall()}


def fetch_all_blocked_with_dns(conn, run_id):
    """
    Fetch ALL blocked domains for a run with full DNS query details:
    domain, status, isp_response, control_response,
    isp_all_ips, isp_ttl, isp_cname, isp_rcode, isp_ms,
    ctrl_all_ips, ctrl_rcode, ctrl_ms, category
    """
    with conn.cursor() as cur:
        cur.execute("""
            SELECT
                mr.domain,
                mr.status,
                mr.isp_response,
                mr.control_response,
                qi.all_responses,
                qi.ttl,
                qi.cname_chain,
                qi.rcode          AS isp_rcode,
                ROUND(qi.response_time_ms::numeric, 2) AS isp_ms,
                qc.all_responses  AS ctrl_all_ips,
                qc.rcode          AS ctrl_rcode,
                ROUND(qc.response_time_ms::numeric, 2) AS ctrl_ms,
                bd.category
            FROM measurement_results mr
            LEFT JOIN dns_queries qi
                ON qi.run_id = mr.run_id
               AND qi.domain = mr.domain
               AND qi.query_type = 'isp'
            LEFT JOIN dns_queries qc
                ON qc.run_id = mr.run_id
               AND qc.domain = mr.domain
               AND qc.query_type = 'control'
            LEFT JOIN blocklist_domains bd ON bd.domain = mr.domain
            WHERE mr.run_id = %s
              AND mr.status LIKE '%%blocked%%'
            ORDER BY mr.domain
        """, (run_id,))
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def fetch_all_results_with_dns(conn, run_id):
    """Fetch ALL results (all statuses) with DNS details."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT
                mr.domain,
                mr.status,
                mr.isp_response,
                mr.control_response,
                qi.all_responses,
                qi.ttl,
                qi.cname_chain,
                qi.rcode          AS isp_rcode,
                ROUND(qi.response_time_ms::numeric, 2) AS isp_ms,
                qc.all_responses  AS ctrl_all_ips,
                qc.rcode          AS ctrl_rcode,
                ROUND(qc.response_time_ms::numeric, 2) AS ctrl_ms,
                bd.category
            FROM measurement_results mr
            LEFT JOIN dns_queries qi
                ON qi.run_id = mr.run_id
               AND qi.domain = mr.domain
               AND qi.query_type = 'isp'
            LEFT JOIN dns_queries qc
                ON qc.run_id = mr.run_id
               AND qc.domain = mr.domain
               AND qc.query_type = 'control'
            LEFT JOIN blocklist_domains bd ON bd.domain = mr.domain
            WHERE mr.run_id = %s
            ORDER BY mr.status, mr.domain
        """, (run_id,))
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def fetch_blocked_sets(conn, run_ids):
    sets = {}
    for rid in run_ids:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT domain FROM measurement_results
                WHERE run_id = %s AND status LIKE '%%blocked%%'
            """, (rid,))
            sets[rid] = {r[0] for r in cur.fetchall()}
    return sets


# ══════════════════════════════════════════════════════════════════════════════
#  Interactive ISP picker
# ══════════════════════════════════════════════════════════════════════════════

def pick_isps(conn):
    rows = fetch_available_isps(conn)
    if not rows:
        print("[ERROR] No completed measurement runs found.")
        sys.exit(1)

    # Latest run per label
    seen = {}
    for r in rows:
        if r[1] not in seen:
            seen[r[1]] = r
    unique = list(seen.values())

    print("\n" + "═" * 66)
    print("  DNS Censorship Analysis — Report Generator  v2")
    print("═" * 66)
    print(f"\n  {len(unique)} ISP measurement run(s) available:\n")
    print(f"  {'#':<4} {'Label':<22} {'Resolver':<18} {'Domains':>8}  {'Date':<17}  ASN")
    print("  " + "─" * 80)
    for i, (rid, label, resolver, started_at, total, asn_org) in enumerate(unique, 1):
        date_str = started_at.strftime("%Y-%m-%d %H:%M") if started_at else "—"
        org = asn_org or ""
        print(f"  {i:<4} {label:<22} {resolver:<18} {(total or 0):>8,}  {date_str:<17}  {org}")
    print()
    print("  Select ISPs to compare (e.g.  1,3  or  1-4  or  all):")
    raw = input("  > ").strip()

    indices = set()
    if raw.lower() == "all":
        indices = set(range(len(unique)))
    else:
        for part in raw.replace(" ", "").split(","):
            if "-" in part:
                a, b = part.split("-", 1)
                indices.update(range(int(a) - 1, int(b)))
            else:
                indices.add(int(part) - 1)

    selected = [unique[i] for i in sorted(indices) if 0 <= i < len(unique)]
    if not selected:
        print("[ERROR] No valid selection.")
        sys.exit(1)
    print(f"\n  ✓ Selected: {', '.join(r[1] for r in selected)}\n")
    return selected


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet: Cover
# ══════════════════════════════════════════════════════════════════════════════

def add_cover(wb, selected, generated_at):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [2, 28, 22, 20, 16, 2])

    title_cell(ws, 1, 2, "DNS CENSORSHIP ANALYSIS", end_col=5, size=20)
    ws.row_dimensions[1].height = 42

    c = ws.cell(row=2, column=2, value="Cross-ISP Comparison Report")
    c.font = Font(bold=False, size=13, color="BFD7FF", name="Calibri", italic=True)
    c.fill = _fill("2563EB")
    c.alignment = _center()
    ws.merge_cells("B2:E2")
    ws.row_dimensions[2].height = 28

    ws.row_dimensions[3].height = 10

    meta = [
        ("Generated",     generated_at.strftime("%Y-%m-%d  %H:%M")),
        ("ISPs Compared", str(len(selected))),
        ("ISP Labels",    "  •  ".join(r[1] for r in selected)),
        ("Control DNS",   "1.1.1.1 (Cloudflare)"),
    ]
    row = 4
    for k, v in meta:
        c1 = ws.cell(row=row, column=2, value=k)
        c1.font = Font(bold=True, size=11, color="1E3A5F", name="Calibri")
        c1.fill = _fill(BG_HEADER_LIGHT)
        c1.alignment = _left()
        c1.border = _border()
        ws.merge_cells(f"C{row}:E{row}")
        c2 = ws.cell(row=row, column=3, value=v)
        c2.font = _font(size=11)
        c2.alignment = _left()
        c2.border = _border()
        ws.row_dimensions[row].height = 20
        row += 1

    ws.row_dimensions[row].height = 12
    row += 1

    # ISP table
    for col, hdr in enumerate(["ISP Label", "Resolver IP", "ASN", "Domains Tested", "Run Date"], 2):
        header_cell(ws, row, col, hdr)
    ws.row_dimensions[row].height = 24
    row += 1

    for i, (rid, label, resolver, started_at, total, asn_org) in enumerate(selected):
        bg = BG_ROW_ALT if i % 2 == 0 else BG_WHITE
        data_cell(ws, row, 2, label,  bg=bg, bold=True, align="left")
        data_cell(ws, row, 3, resolver, bg=bg)
        data_cell(ws, row, 4, asn_org or "—", bg=bg, align="left")
        data_cell(ws, row, 5, total or 0, bg=bg, fmt="#,##0")
        date_str = started_at.strftime("%Y-%m-%d %H:%M") if started_at else "—"
        data_cell(ws, row, 6, date_str, bg=bg)
        ws.row_dimensions[row].height = 20
        row += 1


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet: Summary
# ══════════════════════════════════════════════════════════════════════════════

def add_summary(wb, isp_data):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [28, 14, 13, 13, 13, 13, 11, 12, 12, 11, 14, 14, 14])

    COLS = ["ISP / Label", "Domains\nTested",
            "Blocked\n(IP sig)", "Blocked\n(NXDOM)", "Blocked\n(SERVFAIL)",
            "Total\nBlocked", "Block\nRate %",
            "Accessible", "Changed\nIP", "Timeout",
            "ISP Avg\nLatency ms", "ISP Median\nms", "Control Avg\nms"]

    title_cell(ws, 1, 1, "SUMMARY — Blocking Statistics by ISP", end_col=len(COLS), size=14)
    ws.row_dimensions[1].height = 30

    for ci, h in enumerate(COLS, 1):
        bg = "DC2626" if ci in (3, 4, 5, 6, 7) else BG_HEADER_MID
        header_cell(ws, 2, ci, h, bg=bg, wrap=True)
    ws.row_dimensions[2].height = 36

    for ri, isp in enumerate(isp_data, 3):
        s  = isp["summary"]
        m  = isp["meta"]
        t  = isp.get("timing", {})
        total = m.get("total_domains") or 1
        bl_ip  = s.get("blocked",          0)
        bl_nx  = s.get("blocked_nxdomain", 0)
        bl_sf  = s.get("blocked_servfail", 0)
        total_bl = bl_ip + bl_nx + bl_sf
        acc      = s.get("accessible", 0)
        chg      = s.get("changed",    0)
        tmo      = s.get("timeout",    0)
        pct      = total_bl / total if total else 0
        bg = BG_ROW_ALT if ri % 2 == 0 else BG_WHITE
        pct_bg = BG_RED_LIGHT if pct > 0.10 else (BG_AMBER_LIGHT if pct > 0.01 else BG_GREEN_LIGHT)

        vals = [
            (isp["label"],  bg,     "left",  None,    True),
            (total,         bg,     "right", "#,##0", False),
            (bl_ip,         BG_RED_LIGHT if bl_ip else bg,  "right", "#,##0", False),
            (bl_nx,         BG_RED_LIGHT if bl_nx else bg,  "right", "#,##0", False),
            (bl_sf,         BG_RED_LIGHT if bl_sf else bg,  "right", "#,##0", False),
            (total_bl,      BG_RED_LIGHT if total_bl else bg, "right", "#,##0", True),
            (pct,           pct_bg, "right", "0.00%", True),
            (acc,           bg,     "right", "#,##0", False),
            (chg,           BG_AMBER_LIGHT if chg else bg, "right", "#,##0", False),
            (tmo,           bg,     "right", "#,##0", False),
            (t.get("isp",     {}).get("avg",    "—"), bg, "right", "0.00", False),
            (t.get("isp",     {}).get("median", "—"), bg, "right", "0.00", False),
            (t.get("control", {}).get("avg",    "—"), bg, "right", "0.00", False),
        ]
        for ci, (val, cell_bg, al, fmt, bold) in enumerate(vals, 1):
            data_cell(ws, ri, ci, val, bg=cell_bg, bold=bold, align=al, fmt=fmt)
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = "B3"

    # Bar chart — total blocked
    n = len(isp_data)
    chart_data_col = len(COLS) + 2
    ws.cell(row=2, column=chart_data_col,     value="ISP")
    ws.cell(row=2, column=chart_data_col + 1, value="Total Blocked")
    ws.cell(row=2, column=chart_data_col + 2, value="Accessible")
    for i, isp in enumerate(isp_data, 3):
        s = isp["summary"]
        bl = sum(s.get(k, 0) for k in ["blocked", "blocked_nxdomain", "blocked_servfail"])
        ws.cell(row=i, column=chart_data_col,     value=isp["label"])
        ws.cell(row=i, column=chart_data_col + 1, value=bl)
        ws.cell(row=i, column=chart_data_col + 2, value=s.get("accessible", 0))

    chart = BarChart()
    chart.type     = "col"
    chart.grouping = "clustered"
    chart.title    = "Blocked vs Accessible Domains by ISP"
    chart.y_axis.title = "Domains"
    chart.x_axis.title = "ISP"
    chart.style  = 2
    chart.width  = 22
    chart.height = 14
    data = Reference(ws, min_col=chart_data_col + 1, max_col=chart_data_col + 2,
                     min_row=2, max_row=2 + n)
    cats = Reference(ws, min_col=chart_data_col,     min_row=3, max_row=2 + n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "DC2626"
    chart.series[1].graphicalProperties.solidFill = "16A34A"
    ws.add_chart(chart, f"A{3 + n + 2}")


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet: ISP Detail
# ══════════════════════════════════════════════════════════════════════════════

def add_isp_detail(wb, isp):
    label = isp["label"]
    ws = wb.create_sheet(f"{label[:28]} — Detail")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [28, 40, 2, 22, 22])

    meta = isp["meta"]
    s    = isp["summary"]
    cats = isp["categories"]
    t    = isp.get("timing", {})
    total = meta.get("total_domains") or 1
    total_bl = sum(s.get(k, 0) for k in ["blocked", "blocked_nxdomain", "blocked_servfail"])

    title_cell(ws, 1, 1, f"ISP DETAIL — {label}", end_col=2, size=14)
    ws.row_dimensions[1].height = 30

    row = 3

    # ── Metadata ──────────────────────────────────────────────────────────────
    section_cell(ws, row, 1, "📋  Run Metadata", end_col=2)
    row += 1
    meta_items = [
        ("ISP Label",       meta.get("label", label)),
        ("Resolver IP",     meta.get("resolver_ip", "—")),
        ("Block Signature IP", meta.get("block_ip") or "Not detected"),
        ("ASN",             f"{meta.get('asn','—')}  {meta.get('asn_org','')}".strip()),
        ("Country",         meta.get("country", "—")),
        ("Public IP",       meta.get("public_ip", "—")),
        ("Started At",      str(meta.get("started_at", "—"))[:19]),
        ("Completed At",    str(meta.get("completed_at", "—"))[:19]),
        ("Total Domains",   meta.get("total_domains", 0)),
    ]
    for i, (k, v) in enumerate(meta_items):
        bg = BG_ROW_ALT if i % 2 == 0 else BG_WHITE
        data_cell(ws, row, 1, k,  bg=BG_HEADER_LIGHT, bold=True, align="left")
        data_cell(ws, row, 2, v,  bg=bg, align="left",
                  fmt="#,##0" if isinstance(v, int) else None)
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1

    # ── Status breakdown ──────────────────────────────────────────────────────
    section_cell(ws, row, 1, "🚦  Blocking Status Breakdown", end_col=2)
    row += 1
    header_cell(ws, row, 1, "Status",       bg=BG_HEADER_MID)
    header_cell(ws, row, 2, "Count",        bg=BG_HEADER_MID)
    ws.row_dimensions[row].height = 24
    row += 1

    status_order = ["blocked", "blocked_nxdomain", "blocked_servfail",
                    "accessible", "changed", "timeout"]
    for st in status_order:
        cnt = s.get(st, 0)
        pct = cnt / total if total else 0
        bg  = STATUS_BG.get(st, BG_WHITE)
        lbl = STATUS_LABEL.get(st, st)
        bold = st.startswith("blocked")
        data_cell(ws, row, 1, f"{lbl}  ({pct:.1%})", bg=bg, bold=bold, align="left",
                  fg=FG_RED if bold else FG_DARK)
        data_cell(ws, row, 2, cnt, bg=bg, fmt="#,##0", bold=bold, align="right")
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1

    # ── Timing ────────────────────────────────────────────────────────────────
    section_cell(ws, row, 1, "⏱  DNS Response Timing", end_col=2)
    row += 1
    header_cell(ws, row, 1, "Resolver",    bg=BG_HEADER_MID)
    header_cell(ws, row, 2, "Avg ms  /  Median ms", bg=BG_HEADER_MID)
    ws.row_dimensions[row].height = 24
    row += 1

    for qt, lbl in [("isp", f"ISP  ({meta.get('resolver_ip','')})"),
                     ("control", "Control  (1.1.1.1)")]:
        td = t.get(qt, {})
        avg = td.get("avg", "—")
        med = td.get("median", "—")
        bg  = BG_ROW_ALT if qt == "isp" else BG_WHITE
        data_cell(ws, row, 1, lbl, bg=bg, bold=True, align="left")
        val_str = (f"{avg:.2f} ms  /  {med:.2f} ms"
                   if isinstance(avg, float) else "—")
        data_cell(ws, row, 2, val_str, bg=bg, align="left")
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1

    # ── Category breakdown ────────────────────────────────────────────────────
    if cats:
        section_cell(ws, row, 1, "🗂  Blocked Domains by Category", end_col=2)
        row += 1
        header_cell(ws, row, 1, "Category",       bg=BG_HEADER_MID)
        header_cell(ws, row, 2, "Blocked Count  (% of blocked)", bg=BG_HEADER_MID)
        ws.row_dimensions[row].height = 24
        row += 1
        total_cat = sum(c[1] for c in cats)
        for i, (code, cnt) in enumerate(cats):
            bg  = BG_ROW_ALT if i % 2 == 0 else BG_WHITE
            name = CATEGORY_NAMES.get(code, code)
            pct  = cnt / total_cat if total_cat else 0
            data_cell(ws, row, 1, f"{name}  [{code}]", bg=bg, align="left")
            data_cell(ws, row, 2, f"{cnt:,}  ({pct:.1%})", bg=bg, align="right")
            ws.row_dimensions[row].height = 19
            row += 1

        # Embed category chart PNG
        cat_png = make_category_chart_png(cats, label)
        if cat_png:
            row += 1
            img = XLImage(cat_png)
            img.anchor = f"D{3}"
            ws.add_image(img)


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet: ALL Blocked Domains with DNS details
# ══════════════════════════════════════════════════════════════════════════════

BLOCKED_COLS = [
    ("Domain",                 36, "left"),
    ("Block Type",             20, "center"),
    ("Category",               18, "center"),
    ("ISP Response IP",        18, "center"),
    ("All ISP IPs",            28, "left"),
    ("ISP RCODE",              12, "center"),
    ("ISP TTL (s)",            12, "center"),
    ("ISP Response ms",        14, "center"),
    ("ISP CNAME Chain",        30, "left"),
    ("Control Response IP",    18, "center"),
    ("All Control IPs",        28, "left"),
    ("Control RCODE",          14, "center"),
    ("Control ms",             12, "center"),
]

def add_blocked_domains(wb, isp_label, domains):
    if not domains:
        return
    ws = wb.create_sheet(f"{isp_label[:24]} — Blocked")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [col[1] for col in BLOCKED_COLS])

    title_cell(ws, 1, 1,
               f"ALL BLOCKED DOMAINS — {isp_label}  ({len(domains):,} domains)",
               end_col=len(BLOCKED_COLS), size=13)
    ws.row_dimensions[1].height = 28

    for ci, (hdr, _, _) in enumerate(BLOCKED_COLS, 1):
        header_cell(ws, 2, ci, hdr, wrap=True)
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    for ri, d in enumerate(domains, 3):
        status = d.get("status", "")
        bg = STATUS_BG.get(status, BG_WHITE)
        if ri % 2 == 0 and bg == BG_RED_LIGHT:
            bg = "FECACA"   # slightly darker stripe

        cat_code = d.get("category") or "—"
        cat_name = CATEGORY_NAMES.get(cat_code, cat_code)

        all_isp   = ", ".join(d.get("all_responses")  or []) or d.get("isp_response",  "—")
        all_ctrl  = ", ".join(d.get("ctrl_all_ips")   or []) or d.get("control_response", "—")
        cname     = " → ".join(d.get("cname_chain")   or []) or "—"

        row_vals = [
            (d.get("domain",           ""), "left"),
            (STATUS_LABEL.get(status, status), "center"),
            (cat_name,                         "center"),
            (d.get("isp_response",     "—"),   "center"),
            (all_isp,                          "left"),
            (d.get("isp_rcode",        "—"),   "center"),
            (d.get("ttl"),                     "center"),
            (d.get("isp_ms"),                  "center"),
            (cname,                            "left"),
            (d.get("control_response", "—"),   "center"),
            (all_ctrl,                         "left"),
            (d.get("ctrl_rcode",       "—"),   "center"),
            (d.get("ctrl_ms"),                 "center"),
        ]
        for ci, (val, al) in enumerate(row_vals, 1):
            fmt = "#,##0.00" if ci in (8, 13) else ("#,##0" if ci == 7 else None)
            data_cell(ws, ri, ci, val, bg=bg, align=al, fmt=fmt,
                      size=10, bold=(ci == 1))
        ws.row_dimensions[ri].height = 15


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet: Overlap / Venn
# ══════════════════════════════════════════════════════════════════════════════

def add_overlap_sheet(wb, isp_data, blocked_sets):
    ws = wb.create_sheet("Cross-ISP Overlap")
    ws.sheet_view.showGridLines = False

    labels   = [d["label"] for d in isp_data]
    run_ids  = [d["run_id"] for d in isp_data]
    n        = len(labels)
    set_col_widths(ws, [26] + [16] * n)

    title_cell(ws, 1, 1, "CROSS-ISP OVERLAP — Shared Blocked Domains", end_col=1 + n, size=14)
    ws.row_dimensions[1].height = 28

    # Matrix header
    header_cell(ws, 2, 1, "ISP  ↓  /  ISP  →", bg=BG_HEADER_DARK)
    for ci, lbl in enumerate(labels, 2):
        header_cell(ws, 2, ci, lbl, bg=BG_HEADER_MID)
    ws.row_dimensions[2].height = 28

    for ri, (lbl_a, rid_a) in enumerate(zip(labels, run_ids), 3):
        header_cell(ws, ri, 1, lbl_a, bg=BG_HEADER_MID)
        for ci, (lbl_b, rid_b) in enumerate(zip(labels, run_ids), 2):
            set_a = blocked_sets.get(rid_a, set())
            set_b = blocked_sets.get(rid_b, set())
            if rid_a == rid_b:
                data_cell(ws, ri, ci, len(set_a), bg=BG_HEADER_LIGHT,
                          bold=True, fmt="#,##0")
            else:
                overlap = len(set_a & set_b)
                union   = len(set_a | set_b)
                jaccard = overlap / union * 100 if union else 0
                intensity = int(min(jaccard / 50, 1.0) * 180)
                r_hex = f"FF{255 - intensity:02X}{255 - intensity:02X}"
                c = ws.cell(row=ri, column=ci,
                            value=f"{overlap:,}  ({jaccard:.1f}%)")
                c.font      = _font(size=11)
                c.fill      = _fill(r_hex)
                c.alignment = _center()
                c.border    = _border()
        ws.row_dimensions[ri].height = 24

    # Footnote
    fn_row = 3 + n + 1
    ws.merge_cells(f"A{fn_row}:{get_column_letter(1 + n)}{fn_row}")
    c = ws.cell(row=fn_row, column=1,
                value="Diagonal = total blocked by that ISP.  "
                      "Off-diagonal = shared blocked count (% Jaccard index).  "
                      "Warmer colour = higher overlap.")
    c.font = Font(size=9, italic=True, color="6B7280", name="Calibri")
    c.alignment = _left()
    ws.row_dimensions[fn_row].height = 18

    # ── Venn diagram (2 ISPs only) ─────────────────────────────────────────
    venn_row = fn_row + 2
    ws.merge_cells(f"A{venn_row}:{get_column_letter(1 + n)}{venn_row}")
    sec = ws.cell(row=venn_row, column=1, value="Venn Diagram")
    sec.font  = Font(bold=True, size=12, color="1E3A5F", name="Calibri")
    sec.fill  = _fill(BG_HEADER_LIGHT)
    sec.alignment = _left()
    ws.row_dimensions[venn_row].height = 22

    if len(isp_data) == 2:
        venn_png = make_venn_png(isp_data, blocked_sets)
        if venn_png:
            img = XLImage(venn_png)
            img.width  = 700
            img.height = 410
            img.anchor = f"A{venn_row + 1}"
            ws.add_image(img)
    else:
        # For 3+ ISPs show pairwise text table instead
        note = ws.cell(row=venn_row + 1, column=1,
                       value="(Venn diagram rendered for exactly 2 ISPs; "
                             "see matrix above for multi-ISP overlap.)")
        note.font = Font(size=10, italic=True, color="6B7280", name="Calibri")


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet: Unique Blocks
# ══════════════════════════════════════════════════════════════════════════════

def add_unique_blocks(wb, isp_data, blocked_sets):
    ws = wb.create_sheet("Unique Blocks")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [24, 40, 14])

    run_ids  = [d["run_id"] for d in isp_data]
    all_sets = [blocked_sets.get(rid, set()) for rid in run_ids]

    title_cell(ws, 1, 1,
               "UNIQUE BLOCKS — Domains Blocked by Only One ISP",
               end_col=3, size=14)
    ws.row_dimensions[1].height = 28

    header_cell(ws, 2, 1, "ISP")
    header_cell(ws, 2, 2, "Domain (blocked by this ISP only)")
    header_cell(ws, 2, 3, "Unique Count")
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    row = 3
    for isp, rid, my_set in zip(isp_data, run_ids, all_sets):
        others_union = set().union(*(s for s in all_sets if s is not my_set)) if len(all_sets) > 1 else set()
        unique_set   = my_set - others_union
        unique_list  = sorted(unique_set)

        if not unique_list:
            data_cell(ws, row, 1, isp["label"], bg=BG_HEADER_LIGHT, bold=True, align="left")
            data_cell(ws, row, 2, "(no domains blocked exclusively by this ISP)",
                      bg=BG_GREY_LIGHT, fg=FG_MUTED, align="left")
            data_cell(ws, row, 3, 0, bg=BG_GREY_LIGHT, fmt="#,##0")
            ws.row_dimensions[row].height = 20
            row += 1
        else:
            for i, domain in enumerate(unique_list):
                bg = BG_ROW_ALT if i % 2 == 0 else BG_WHITE
                if i == 0:
                    data_cell(ws, row, 1, isp["label"],      bg=BG_HEADER_LIGHT, bold=True, align="left")
                    data_cell(ws, row, 3, len(unique_list),  bg=BG_HEADER_LIGHT, bold=True, fmt="#,##0")
                else:
                    data_cell(ws, row, 1, "", bg=BG_HEADER_LIGHT)
                    data_cell(ws, row, 3, "", bg=BG_HEADER_LIGHT)
                data_cell(ws, row, 2, domain, bg=bg, align="left", size=10)
                ws.row_dimensions[row].height = 15
                row += 1

        # separator row
        for ci in range(1, 4):
            c = ws.cell(row=row, column=ci, value="")
            c.fill  = _fill(BG_HEADER_DARK)
            c.border = _border()
        ws.row_dimensions[row].height = 4
        row += 1


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Generate ISP DNS censorship comparison report (.xlsx)")
    parser.add_argument("--db",     default=os.environ.get("DATABASE_URL"))
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    if not args.db:
        print("[ERROR] No DATABASE_URL. Set it in .env or pass --db.")
        sys.exit(1)

    print("[INFO] Connecting to database...")
    try:
        conn = db_connect(args.db)
        print("[INFO] Connected ✓")
    except Exception as e:
        print(f"[ERROR] {e}")
        sys.exit(1)

    selected = pick_isps(conn)

    print("[INFO] Fetching data...")
    isp_data = []
    for (rid, label, resolver, started_at, total, asn_org) in selected:
        print(f"  ↳ {label}  (run_id={rid})  — fetching all blocked domains + DNS details...")
        isp_data.append({
            "run_id":     rid,
            "label":      label,
            "meta":       fetch_run_meta(conn, rid),
            "summary":    fetch_run_summary(conn, rid),
            "categories": fetch_category_breakdown(conn, rid),
            "timing":     fetch_avg_response_times(conn, rid),
            "domains":    fetch_all_blocked_with_dns(conn, rid),
        })
        print(f"     {len(isp_data[-1]['domains']):,} blocked domains loaded.")

    print("[INFO] Computing cross-ISP overlap...")
    run_ids      = [d["run_id"] for d in isp_data]
    blocked_sets = fetch_blocked_sets(conn, run_ids)
    conn.close()

    print("[INFO] Building Excel workbook...")
    wb            = Workbook()
    generated_at  = datetime.now()

    add_cover(wb, selected, generated_at)
    add_summary(wb, isp_data)
    for isp in isp_data:
        add_isp_detail(wb, isp)
        add_blocked_domains(wb, isp["label"], isp["domains"])
    add_overlap_sheet(wb, isp_data, blocked_sets)
    add_unique_blocks(wb, isp_data, blocked_sets)

    # Output path
    if args.output:
        out = Path(args.output)
    else:
        ts    = generated_at.strftime("%Y%m%d_%H%M%S")
        names = "_vs_".join(d["label"][:10].replace(" ", "") for d in isp_data[:4])
        out   = Path(f"isp_report_{names}_{ts}.xlsx")

    wb.save(out)
    print(f"\n{'═' * 58}")
    print(f"  ✓  Report saved  →  {out}")
    print(f"  Sheets  : {', '.join(wb.sheetnames)}")
    print(f"  ISPs    : {len(isp_data)}")
    total_domains = sum(len(d["domains"]) for d in isp_data)
    print(f"  Domains : {total_domains:,} blocked records (all, no cap)")
    print(f"{'═' * 58}\n")


if __name__ == "__main__":
    main()